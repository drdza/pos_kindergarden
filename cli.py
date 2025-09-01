# cli.py
from datetime import datetime
import typer, shutil
from pathlib import Path
import sqlite3
from config import DB_PATH
from db.init import init_db, schema_path, get_connection
from db.migrate import apply_migrations
from typing import Optional
import math
try:
    import openpyxl  # para .xlsx
except Exception:
    openpyxl = None

app = typer.Typer(help="Herramientas de base de datos (SQLite)")

@app.command("init-db")
def init_db_cmd():
    """Crea/reescribe la estructura base desde db/schema.sql"""
    typer.echo(f"DB: {DB_PATH}")
    init_db(schema_path())
    typer.secho("OK: esquema inicial aplicado.", fg=typer.colors.GREEN)

@app.command("migrate")
def migrate_cmd():
    """Aplica migraciones en db/migrations/*.sql"""
    done = apply_migrations(Path("db") / "migrations")
    if done:
        for f in done: typer.echo(f"aplicada: {f}")
    else:
        typer.echo("No hay migraciones pendientes.")

@app.command("seed")
def seed_cmd(seed_file: Path = typer.Option(Path("db")/"seed.sql")):
    """Ejecuta datos semilla"""
    if not seed_file.exists():
        typer.echo("Seed no encontrado.")
        raise typer.Exit(code=1)
    sql = seed_file.read_text(encoding="utf-8")
    with get_connection() as conn:
        conn.executescript(sql)
        conn.commit()
    typer.secho("OK: seed aplicado.", fg=typer.colors.GREEN)

@app.command("integrity-check")
def integrity_check():
    """PRAGMA integrity_check"""
    with get_connection() as conn:
        res = conn.execute("PRAGMA integrity_check;").fetchone()[0]
    typer.secho(f"integrity_check: {res}", fg=typer.colors.GREEN)

@app.command("vacuum")
def vacuum():
    """Compacción y mantenimiento"""
    with get_connection() as conn:
        conn.execute("VACUUM;")
        conn.commit()
    typer.secho("OK: VACUUM.", fg=typer.colors.GREEN)

@app.command("backup")
def backup(dst: Optional[Path] = typer.Option(  # <-- Optional[Path]
    None,
    help="Ruta del archivo destino. Si se omite, se crea data/backup_YYYY-MM-DD.db",
)):
    """Copia en caliente usando backup API de sqlite3"""

    if dst is None:
        dst = Path("data") / f"backup_{datetime.now():%Y-%m-%d}.db"

    dst.parent.mkdir(exist_ok=True)
    
    with get_connection() as src_conn, sqlite3.connect(dst) as dst_conn:
        src_conn.backup(dst_conn)
    typer.secho(f"Backup creado en: {dst}", fg=typer.colors.GREEN)


@app.command("import-customers")
def import_customers_xlsx(
    src: Path = typer.Argument(..., help="Ruta al Excel Alumnos.xlsx (hoja ControlUnico)"),
    sheet: str = typer.Option("ControlUnico", help="Nombre de la hoja"),
    dry_run: bool = typer.Option(False, help="Solo mostrar acciones; no escribe en BD"),
):
    """
    Importa/actualiza alumnos (customers) a partir del Excel real que compartiste.
    Upsert por 'enrollment' (MATRICULA). Crea grades/groups/shifts por 'code' si faltan.
    """
    if openpyxl is None:
        typer.secho("Falta openpyxl. Instala con: pip install openpyxl", fg="red")
        raise typer.Exit(code=1)
    if not src.exists():
        typer.secho(f"No existe: {src}", fg="red"); raise typer.Exit(code=1)

    wb = openpyxl.load_workbook(src, data_only=True)
    if sheet not in wb.sheetnames:
        typer.secho(f"La hoja '{sheet}' no existe. Hojas: {wb.sheetnames}", fg="red"); raise typer.Exit(1)
    ws = wb[sheet]

    # Header map: nombre exacto de tus columnas → clave interna
    header_map = {
        "MATRICULA": "enrollment",
        "NOMBRE": "first_name",
        "APELLIDO PATERNO": "ap_paterno",
        "APELLIDO MATERNO": "ap_materno",
        "DOMICILIO": "address",
        "GRADO": "grade_code",
        "GRUPO": "group_code",
        "HORARIO": "shift_text",     # MATUTINO / VESPERTINO
        "SEXO": "gender",
        "NOMBRE MAMÁ": "fullname_mom",
        "NOMBRE PAPÁ": "fullname_dad",
        "DIA FNAC": "birth_day",
        "MES FNAC": "birth_month_es",
        "AÑO FNAC": "birth_year",
        "CURP": "curp",
        "CELULAR 1": "phone",
        "CELULAR 2": "mobile_phone",
        "REFERENCIA": "pay_reference",
    }

    # Lee encabezados
    headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column+1)]
    idx = {}
    for k_excel, k_int in header_map.items():
        try:
            idx[k_int] = headers.index(k_excel) + 1
        except ValueError:
            idx[k_int] = None  # columna opcional ausente

    def cell(row, key):
        col = idx.get(key)
        if not col: return None
        v = ws.cell(row=row, column=col).value
        if v is None: return None
        if isinstance(v, str): return v.strip()
        return v

    # Normalizadores
    MONTHS_ES = {
        "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
        "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "SETIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12
    }
    def to_birth_date(day, month_es, year):
        if day in (None, "", 0) or year in (None, "", 0) or not month_es:
            return None
        try:
            d = int(day) if not isinstance(day, float) else int(day)
            y = int(year) if not isinstance(year, float) else int(year)
            if isinstance(month_es, (int, float)) and not math.isnan(month_es):
                m = int(month_es)
            else:
                m = MONTHS_ES.get(str(month_es).strip().upper())
            if not m: return None
            return datetime(y, m, d).strftime("%Y-%m-%d")
        except Exception:
            return None

    def to_shift_code(txt):
        if not txt: return None
        t = str(txt).strip().upper()
        if t.startswith("MAT"): return "MAT"
        if t.startswith("VES"): return "VES"
        return None

    def to_gender(g):
        if not g: return None
        gg = str(g).strip().upper()
        if gg.startswith("M"): return "F"
        if gg.startswith("H"): return "M"
        return None

    def to_text(x):
        if x is None: return None
        s = str(x).strip()
        return s if s != "" else None

    # Conexión
    conn = get_connection()
    cur = conn.cursor()

    # Helpers: obtener/crear IDs de catálogos
    def get_or_create(table, code) -> Optional[int]:
        if not code: return None
        r = cur.execute(f"SELECT id FROM {table} WHERE code=?", (code,)).fetchone()
        if r: return r[0]
        cur.execute(f"INSERT INTO {table}(code, name) VALUES(?, ?)", (code, code))
        return cur.lastrowid

    inserts = updates = errors = 0

    for r in range(2, ws.max_row+1):
        try:
            enrollment = cell(r, "enrollment")
            # enrollment podría venir como número → conviértelo a texto sin decimales
            if enrollment is None or str(enrollment).strip() == "":
                # fila vacía, continúa
                continue
            if isinstance(enrollment, float):
                enrollment = str(int(enrollment))
            else:
                enrollment = str(enrollment).strip()

            first_name = to_text(cell(r, "first_name"))
            ap1 = to_text(cell(r, "ap_paterno"))
            ap2 = to_text(cell(r, "ap_materno"))
            second_name = " ".join([x for x in [ap1, ap2] if x]).strip() or None

            address = to_text(cell(r, "address"))
            grade_code = to_text(cell(r, "grade_code"))
            group_code = to_text(cell(r, "group_code"))
            shift_text = to_text(cell(r, "shift_text"))
            shift_code = to_shift_code(shift_text)

            gender = to_gender(cell(r, "gender"))
            mom = to_text(cell(r, "fullname_mom"))
            dad = to_text(cell(r, "fullname_dad"))

            birth_day = cell(r, "birth_day")
            birth_month_es = cell(r, "birth_month_es")
            birth_year = cell(r, "birth_year")
            birth_date = to_birth_date(birth_day, birth_month_es, birth_year)

            phone = cell(r,"phone")
            mobile_phone = cell (r, "mobile_phone")

            curp = to_text(cell(r, "curp"))
            pay_ref = to_text(cell(r, "pay_reference"))

            # Resuelve catálogos (crea si no existen)
            grade_id = get_or_create("grades", grade_code) if grade_code else None
            group_id = get_or_create("groups", group_code) if group_code else None
            shift_id = get_or_create("shifts", shift_code) if shift_code else None

            fields = {
                "enrollment": enrollment,
                "first_name": first_name,
                "second_name": second_name,     # apellidos concatenados
                "address": address,
                "grade_id": grade_id,
                "group_id": group_id,
                "shift_id": shift_id,
                "gender": gender,
                "fullname_mom": mom,
                "fullname_dad": dad,
                "birth_date": birth_date,
                "curp": curp,
                "phone": phone,
                "mobile_phone": mobile_phone,
                "pay_reference": pay_ref,
                "active": 1,
            }

            # Validación mínima
            if not first_name:
                raise ValueError("NOMBRE vacío")
            # Upsert por enrollment
            exists = cur.execute("SELECT id FROM customers WHERE enrollment=?", (enrollment,)).fetchone()
            if dry_run:
                action = "UPDATE" if exists else "INSERT"
                typer.echo(f"[fila {r}] {action} {enrollment} → {fields}")
            else:
                if exists:
                    sets = ", ".join([f"{k}=?" for k in fields.keys() if k != "enrollment"])
                    params = [fields[k] for k in fields.keys() if k != "enrollment"] + [enrollment]
                    cur.execute(f"UPDATE customers SET {sets} WHERE enrollment=?", params)
                    updates += 1
                else:
                    cols = ",".join(fields.keys())
                    qs = ",".join(["?"]*len(fields))
                    cur.execute(f"INSERT INTO customers({cols}) VALUES({qs})", tuple(fields.values()))
                    inserts += 1

        except Exception as e:
            errors += 1
            typer.secho(f"[fila {r}] ERROR: {e}", fg="red")

    if dry_run:
        conn.rollback()
        typer.secho(f"Dry-run: inserts={inserts}, updates={updates}, errores={errors}", fg="yellow")
    else:
        conn.commit()
        typer.secho(f"Import OK: inserts={inserts}, updates={updates}, errores={errors}", fg="green")


@app.command("test-sale")
def test_sale():
    from services.sales_service import alta_venta

    # Usa SKUs que existan en tu seed. Si en tu seed el SKU es "P-001", ajusta aquí.
    result = alta_venta(
        customer_name="Juan Pérez",
        seller_name="Caja Mostrador",
        customer_id=1,   # opcional si ya tienes ese alumno; si no, quita este parámetro
        seller_id=1,     # idem
        items_ui=[
            {"sku":"P-001", "descripcion":"Servicio 1", "qty":1, "unit_price":25.0, "tax_rate":0.00},
            {"descripcion":"Servicio 2", "qty":1, "unit_price":50.0, "tax_rate":0.00},  # sin SKU, captura libre
        ],
        payments=[{"method":"cash", "amount":75.0}]  # method es libre: 'cash'/'card'/'transfer'
    )
    print("Venta creada ->", result)   # {'id': 1, 'folio': 'F0001'}


@app.command("show-sale")
def show_sale(
    sale_id: int = typer.Option(..., "--sale_id", "-s", help="ID numérico de la venta")
):
    from repos.sales import get_sale
    sale, items, pays = get_sale(sale_id)
    print("SALE:", sale)
    print("ITEMS:", items)
    print("PAYMENTS:", pays)




if __name__ == "__main__":
    app()
